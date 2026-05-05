"""
Parser principal: le a planilha de FEDERAIS e gera convenios.json estruturado.

Estrutura igual ao gov/pref:
  - linha 1: nome do convenio (col 2+) — opcional
  - linha 2: bancos (col 2+) - pode ter "**SUSPENSO**" no nome
  - linhas 3+: cada linha = 1 atributo (col 1 = nome do atributo, cols 2+ = valor por banco)
  - linhas com so col 1 preenchida = secao (PORTABILIDADE, CARTAO, PUBLICO ALVO)

DIFERENCAS pra GOV/PREF:
  - Convenios federais nao tem UF (sao nacionais)
  - Adiciona campos `categoria` (civil|militar) e `orgao` (SIAPE, SERPRO, EXERCITO, MARINHA, AERONAUTICA)
  - SIAPE e quebrado em 4 sub-convenios (1 por tipo de operacao):
      SIAPE NOVO-REFIN, SIAPE PORTABILIDADE, CARTAO SIAPE, CARTAO BENEFICIO
  - Outros convenios (SERPRO, MARINHA, EXERCITO, AERONAUTICA) sao 1 aba = 1 convenio completo

Saida: scripts/fed/convenios.json
"""
import json, re, sys, unicodedata
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(__file__).parent.parent / 'FEDERAIS_RESUMO.xlsx'
OUT = Path(__file__).parent / 'convenios.json'

# ── Mapa: aba -> {nome amigavel, categoria, orgao, operacao_tipo} ──
SHEET_MAP = {
    'SIAPE NOVO-REFIN': {
        'nome': 'SIAPE - Novo / Refinanciamento',
        'categoria': 'civil',
        'orgao': 'SIAPE',
        'operacao_tipo': 'novo_refin',
    },
    'SIAPE PORTABILIDADE': {
        'nome': 'SIAPE - Portabilidade',
        'categoria': 'civil',
        'orgao': 'SIAPE',
        'operacao_tipo': 'portabilidade',
    },
    'CARTÃO SIAPE': {
        'nome': 'SIAPE - Cartão Consignado (RMC)',
        'categoria': 'civil',
        'orgao': 'SIAPE',
        'operacao_tipo': 'cartao_consignado',
    },
    'CARTÃO BENEFICIO': {
        'nome': 'SIAPE - Cartão Benefício (RCC)',
        'categoria': 'civil',
        'orgao': 'SIAPE',
        'operacao_tipo': 'cartao_beneficio',
    },
    'SERPRO': {
        'nome': 'SERPRO',
        'categoria': 'civil',
        'orgao': 'SERPRO',
        'operacao_tipo': 'completo',
    },
    'AERONÁUTICA': {
        'nome': 'Forças Armadas - Aeronáutica',
        'categoria': 'militar',
        'orgao': 'AERONAUTICA',
        'operacao_tipo': 'completo',
    },
    'EXÉRCITO': {
        'nome': 'Forças Armadas - Exército',
        'categoria': 'militar',
        'orgao': 'EXERCITO',
        'operacao_tipo': 'completo',
    },
    'MARINHA': {
        'nome': 'Forças Armadas - Marinha',
        'categoria': 'militar',
        'orgao': 'MARINHA',
        'operacao_tipo': 'completo',
    },
}

# Abas que sao indices/separadores/auxiliares — pular
SKIP_SHEETS = {
    'Planilha1','Planilha2','FEDERAL','SIAPE','UPAGS COM MARGEM DE SEGURANÇA',
    'SIAPE PUBLICO ALVO','SIAPE LIMITE OPERACIONAL','FORÇAS ARMADAS',
    'PREC-CP','PREC-CP ','PREC - FUTURO PREV',
    # Abas UPAGs (listagens auxiliares por banco — usadas pelo cruzamento, nao sao convenios)
    'V8 UPAGs','MERCANTIL UPAGs','QUERO MAIS UPAGs','C6 UPAGs','DAY UPAGs',
    'PRESENÇABANK UPAGS','FACTA UPAGS ACEITAS','PAN UPAGs','SABEMI UPAGs',
    'SAFRA UPAGs','FUTURO UPAGs','BRB UPAGs','PHTECH UPAGs','HAPPY UPAGs',
}

def slugify(s):
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^a-zA-Z0-9]+', '-', s).strip('-').lower()
    return s

def clean_text(s):
    n = unicodedata.normalize('NFKD', str(s))
    n = ''.join(c for c in n if not unicodedata.combining(c))
    return n.strip()

def is_suspenso(banco_name):
    if not banco_name: return False
    n = clean_text(banco_name).upper()
    return ('SUSPENSO' in n or 'SUPSENSO' in n or 'BLOQUEADO' in n
            or 'INATIVO' in n or 'TEMPORARIAMENTE' in n)

def clean_banco_name(n):
    if not n: return ''
    s = str(n)
    s = s.split('\n')[0].split('\r')[0].strip()
    s = re.sub(r'\*+[^*]*\*+', '', s)
    s = re.sub(r'\bTEMPORARIAMENTE\s+SUSPENS\w*\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUSPENSO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUPSENSO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bBLOQUEADO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bINATIVO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bTEMPORARIAMENTE\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUSPENS\w*\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bREFIN\b', '', s, flags=re.IGNORECASE)
    s = s.replace('**','').strip(' -:*')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def is_valid_banco_name(n):
    if not n: return False
    s = str(n).strip()
    primeira_linha = s.split('\n')[0].split('\r')[0].strip()
    if len(primeira_linha) > 60: return False
    if len(primeira_linha) < 2: return False
    if ':' in primeira_linha and 'SUSPEN' not in primeira_linha.upper(): return False
    bad_prefixes = ('apenas','havendo','somente','obs','observacao','para ','quando ',
                    'os mesmos','dever','a partir','reconhecimento','politicas',
                    'política','politica')
    if primeira_linha.lower().startswith(bad_prefixes): return False
    if not re.search(r'[a-zA-Z]', primeira_linha): return False
    return True

# ── Mapeamento de atributos canonicos ─────────────────────────
ATTR_MAP = [
    ('operacoes',                ['quais operacoes', 'operacoes realizadas', 'operacoes que realiza']),
    ('valor_minimo',             ['valor minimo de operacao', 'valor minimo', 'valor minimo de operacao novo']),
    ('margem_utilizavel',        ['margem utilizavel']),
    ('data_corte',               ['data de corte', 'virada de folha']),
    ('qtd_contratos',            ['qtde.* de contrato', 'quantidade.*contrato', 'quantidade de averbacao']),
    ('margem_seguranca',         ['^margem de seguranca$']),
    ('margem_seguranca_refin',   ['margem de seguranca p/ refin','margem de seguranca refin']),
    ('min_parcelas_refin',       ['minimo de parcelas pagas para refinanciar', '% de pagto de contrato para refin']),
    ('pode_agregar_margem',      ['pode agregar margem no refin', 'pode agregar margem', 'agregar margem refin']),
    ('reserva_margem',           ['^reserva de margem', 'tipo de averbacao']),
    ('pode_abater_negativo',     ['pode abater o negativo', 'margem negativa']),
    ('pode_unificar_parcelas',   ['pode unificar as parcelas', 'unificar parcelas refin']),
    ('formalizacao_digital',     ['faz formalizacao digital', 'formalizacao digital']),
    ('limite_idade',             ['limite de idade', 'idade x valor', 'politica de idade', 'multiplicador x limite']),
    ('limite_credito',           ['limite de credito', 'limite operacional']),
    ('liberacao_credito',        ['liberacao do credito', 'liberacao do credito']),
    ('reducao_parcela',          ['reducao de parcela']),
    ('compra_divida',            ['compra de divida']),
    ('valor_troco_refin',        ['valor minimo troco refin']),
    ('outras_informacoes',       ['outras informacoes']),
    ('regras_residencia',        ['regras para clientes que residem']),
    ('port_saldo_minimo',        ['saldo minimo', 'saldo minimo portabilidade']),
    ('port_min_parcelas_pagas',  ['minimo de parcelas pagas']),
    ('port_taxa_minima',         ['taxa minima']),
    ('port_parcela_minima',      ['parcela minima']),
    ('port_autorizacao_especial',['precisa de autorizacao especial', 'precisa de autorizacao']),
    ('port_pagamento_saldo',     ['pagamento de saldo']),
    ('port_doc_pagamento_saldo', ['documentacao obrigatoria para pagamento']),
    ('port_analise_quando',      ['analise antes ou depois do saldo']),
    ('port_reduz_parcela',       ['faz reducao de parcela', 'reducao de parcela na port']),
    ('port_agrega_margem',       ['^agrega margem', 'agregar margem refin portabilidade']),
    ('port_unifica_parcelas',    ['^unifica parcelas', 'unifica parcelas na port']),
    ('port_risco',               ['risco banco ou correspondente', 'tabela port garantida risco', 'risco operacao refin', 'tabela de risco']),
    ('port_compra_divida',       ['compra de divida']),
    ('port_o_que_precisa_compra',['o que precisa para realizar a compra']),
    ('port_atuacao_saldo',       ['atuacao saldo']),
    ('port_rco_tip',             ['rco / tip', 'rco/tip', 'rco tip']),
    ('port_bancos_nao_porta',    ['bancos que nao porta', 'bancos nao porta']),
    ('port_troco_minimo',        ['troco minimo refin de port', 'troco minimo refin port']),
    ('port_pura',                ['trabalha com portabilidade pura']),
    ('port_fase_retorno',        ['fase / atividade no retorno']),
    ('port_documentacao',        ['documentacao obrigatoria para solic']),
    ('port_caracteristicas',     ['caracteristicas especificas do conv']),
    ('port_emite_comprovante',   ['como emitir o comprovante de pagame']),
    ('port_ajuste_parcela',      ['faz ajuste de parcela']),
    ('port_digitacao_refin',     ['digitacao do refin de port']),
    ('cb_limite_minimo',         ['limite minimo de credito']),
    ('cb_valor_minimo_saque',    ['valor minimo de saque']),
    ('cb_valor_minimo_margem',   ['valor minimo de margem']),
    ('cb_limite_maximo_saque',   ['limite maximo de saque']),
    ('cb_taxa_juros',            ['^taxa de juros$']),
    ('cb_seguro',                ['valor seguro prestamista']),
    ('cb_desconto_seguro',       ['desconto do seguro', 'desconto de seguro']),
    ('cb_analfabeto',            ['analfabeto']),
    ('cb_aumento_limite',        ['aumento de limite', 'faz aumento de limite']),
    ('cb_tarifa_emissao',        ['tarifa de emissao do cartao']),
    ('cb_bandeira',              ['bandeira do cartao']),
    ('cb_valor_maximo_op',       ['valor maximo de op']),
    ('cb_margem',                ['margem do cartao']),
    ('cb_relacao_upags',         ['relacao de upags']),
    ('cb_publico_atendido',      ['publico atendido', 'publico atendido']),
    ('cb_publico_nao_atendido',  ['publico nao atendido']),
    ('cb_faz_refin',             ['faz refin']),
]

SECTION_HEADERS = ['portabilidade','cartao','cartão','publico alvo','público alvo',
                   'politicas exercito','política exército','politica exercito']

# Siglas/primeiras palavras de bancos conhecidos (usado pra identificar linha de banco
# quando aparece um banco unico, evitando confundir com titulos como AERONAUTICA/SERPRO)
BANCOS_CONHECIDOS = {
    'PAN','SAFRA','ALFA','ITAU','ITAÚ','DAYCOVAL','CCB','BANRISUL','BMG','CREFISA',
    'C6','SABEMI','FACTA','FUTURO','BRB','PH','PHTECH','PRESENCA','PRESENÇA','AKI',
    'HOPE','BTW','QUERO','PARANA','PARANÁ','BRADESCO','SANTANDER','AGIBANK','AMIGOZ',
    'OLE','OLÉ','MASTER','MERCANTIL','V8','MEU','CAPITAL','INTER','BV','SICREDI','SICOOB',
    'CAIXA','BB','OUROCAP','GRANSAFRA','PINE','PICPAY','ZEMA','MAXIMA','MÁXIMA',
}

def normalize_attr_name(s):
    if not s: return ''
    n = unicodedata.normalize('NFKD', str(s))
    n = ''.join(c for c in n if not unicodedata.combining(c))
    n = re.sub(r'\s+', ' ', n).strip().lower()
    return n

def match_attr_slug(attr_name):
    n = normalize_attr_name(attr_name)
    for slug, patterns in ATTR_MAP:
        for p in patterns:
            if re.search(p, n):
                return slug
    return None

def is_section_header(attr_name):
    n = normalize_attr_name(attr_name)
    return n in SECTION_HEADERS or any(n.startswith(s) for s in SECTION_HEADERS)

def parse_idade_range(text):
    if not text: return None, None
    s = str(text)
    m = re.search(r'(\d{1,2})\s*(?:a|à|até|completos)?\s*(\d{2})\s*anos', s, re.IGNORECASE)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 16 <= a <= 99 and 16 <= b <= 99 and a < b:
            return a, b
    return None, None

def parse_margem(text):
    if text is None: return None
    s = str(text).strip()
    try:
        f = float(s.replace(',','.'))
        if 0 < f <= 1: return round(f,4)
        if 1 < f <= 100: return round(f/100,4)
    except: pass
    m = re.search(r'(\d+[\.,]?\d*)\s*%', s)
    if m:
        try: return round(float(m.group(1).replace(',','.'))/100, 4)
        except: pass
    return None

def parse_taxa(text):
    if text is None: return None
    s = str(text).strip()
    try:
        f = float(s.replace(',','.'))
        if 0 < f < 0.1: return round(f,4)
        if 0.5 < f < 5: return round(f/100,4)
    except: pass
    m = re.search(r'(\d+[\.,]\d+)\s*%?', s)
    if m:
        try:
            f = float(m.group(1).replace(',','.'))
            if 0 < f < 0.1: return round(f,4)
            if 0.5 < f < 5: return round(f/100,4)
        except: pass
    return None

def parse_operacoes(text, sheet_meta):
    """Detecta operacoes operadas. Se a aba ja eh especifica, marca conforme."""
    op_tipo = sheet_meta.get('operacao_tipo','completo')
    base = {'novo':False,'refin':False,'port':False,'cartao':False}
    if op_tipo == 'novo_refin':
        base = {'novo':True,'refin':True,'port':False,'cartao':False}
    elif op_tipo == 'portabilidade':
        base = {'novo':False,'refin':False,'port':True,'cartao':False}
    elif op_tipo in ('cartao_consignado','cartao_beneficio'):
        base = {'novo':False,'refin':False,'port':False,'cartao':True}
    if not text:
        return base
    s = str(text).upper()
    return {
        'novo':   base['novo']   or bool(re.search(r'\bNOVO\b', s)),
        'refin':  base['refin']  or bool(re.search(r'\bREFIN', s)),
        'port':   base['port']   or bool(re.search(r'\bPORT', s)),
        'cartao': base['cartao'] or bool(re.search(r'CART[ÃA]O|\bCB\b|\bRMC\b|\bRCC\b', s)),
    }

# ─── PARSE ────────────────────────────────────────────────────
print(f'Lendo {XLSX.name}...')
wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
convenios = []
problemas = []

for sn in wb.sheetnames:
    sn_clean = sn.strip()
    if sn_clean in SKIP_SHEETS:
        continue
    if sn_clean not in SHEET_MAP:
        problemas.append({'sheet': sn, 'motivo': 'aba nao mapeada (nao eh convenio principal)'})
        continue

    meta = SHEET_MAP[sn_clean]
    ws = wb[sn]

    # max_row pode vir gigante (1048575 — todo Excel com formula); limita
    max_rows = min(ws.max_row or 0, 100)
    max_cols = min(ws.max_column or 1, 30)
    if max_cols < 2:
        problemas.append({'sheet': sn, 'motivo': 'sem colunas'})
        continue

    rows = []
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        rows.append([v for v in row])
    rows = [r for r in rows if any(v not in (None, '') and str(v).strip() for v in r)]
    if len(rows) < 3:
        problemas.append({'sheet': sn, 'motivo': f'so {len(rows)} linhas com dados'})
        continue

    # Linha 1: nome do convenio (na planilha) — usamos o nome amigavel do mapa
    # Linha 2 (idx 1): bancos.
    # detecta linha de bancos: linha que tenha >= 2 candidatos OU 1 candidato com sigla de banco conhecido
    bancos_header = None
    bancos_row_idx = None
    for ridx in range(min(3, len(rows))):
        row = rows[ridx]
        candidates = []
        for i, v in enumerate(row[1:], start=1):
            if v and str(v).strip() and is_valid_banco_name(v):
                candidates.append({'col': i, 'nome_raw': str(v).strip()})
        if len(candidates) >= 2:
            bancos_header = candidates
            bancos_row_idx = ridx
            break
        if len(candidates) == 1:
            # Confere se eh um banco conhecido (nao um titulo tipo SERPRO/AERONAUTICA/CARTAO BENEFICIO)
            first_word = clean_text(candidates[0]['nome_raw']).upper().split()[0] if candidates[0]['nome_raw'] else ''
            if first_word in BANCOS_CONHECIDOS:
                bancos_header = candidates
                bancos_row_idx = ridx
                break
    if not bancos_header:
        problemas.append({'sheet': sn, 'motivo': 'sem linha de bancos identificada'})
        continue

    # Linhas pos-bancos: atributos
    atributos = []
    secao_atual = 'principal'
    for r in rows[bancos_row_idx + 1:]:
        attr_label = r[0]
        if not attr_label or not str(attr_label).strip(): continue
        attr_label = str(attr_label).strip()
        valores_cols = {i: ('' if v is None else str(v).strip()) for i, v in enumerate(r[1:], start=1)}
        if is_section_header(attr_label) and not any(valores_cols.values()):
            n = normalize_attr_name(attr_label)
            if 'portabilidade' in n: secao_atual = 'portabilidade'
            elif 'cartao' in n or 'cartão' in n: secao_atual = 'cartao'
            elif 'publico' in n or 'público' in n: secao_atual = 'publico_alvo'
            elif 'exercito' in n or 'exército' in n or 'politicas' in n: secao_atual = 'principal'
            continue
        slug = match_attr_slug(attr_label)
        atributos.append({
            'slug': slug, 'label_raw': attr_label,
            'secao': secao_atual, 'valores': valores_cols
        })

    bancos_out = []
    for b in bancos_header:
        col = b['col']
        nome_raw = b['nome_raw']
        suspenso = is_suspenso(nome_raw)
        nome_clean = clean_banco_name(nome_raw)
        if not nome_clean or len(nome_clean) < 2: continue

        atrib_canonicos = {}
        atrib_brutos = []
        for a in atributos:
            valor = a['valores'].get(col, '').strip()
            if not valor: continue
            atrib_brutos.append({
                'label': a['label_raw'], 'valor': valor, 'secao': a['secao']
            })
            if a['slug']:
                atrib_canonicos[a['slug']] = valor

        operacoes = parse_operacoes(atrib_canonicos.get('operacoes',''), meta)
        idade_min, idade_max = parse_idade_range(
            atrib_canonicos.get('limite_idade','') + ' ' + atrib_canonicos.get('limite_credito','')
        )
        margem = parse_margem(atrib_canonicos.get('margem_utilizavel'))
        taxa = parse_taxa(atrib_canonicos.get('port_taxa_minima'))

        bancos_out.append({
            'nome': nome_clean,
            'nome_raw': nome_raw,
            'slug': slugify(nome_clean),
            'suspenso': suspenso,
            'operacoes': operacoes,
            'idade_min': idade_min,
            'idade_max': idade_max,
            'margem_utilizavel': margem,
            'taxa_minima_port': taxa,
            'atributos': atrib_canonicos,
            'atributos_brutos': atrib_brutos,
        })

    if not bancos_out:
        problemas.append({'sheet': sn, 'motivo': 'sem bancos validos apos limpeza'})
        continue

    convenios.append({
        'sheet': sn,
        'slug': slugify(meta['nome']),
        'nome': meta['nome'],
        'categoria': meta['categoria'],
        'orgao': meta['orgao'],
        'operacao_tipo': meta['operacao_tipo'],
        'qtd_bancos': len(bancos_out),
        'qtd_atributos': len(atributos),
        'bancos': bancos_out,
    })

# ─── Agrega bancos unicos ─────────────────────────────────────
bancos_unicos = {}
for c in convenios:
    for b in c['bancos']:
        slug = b['slug']
        if slug not in bancos_unicos:
            bancos_unicos[slug] = {'slug': slug, 'nome': b['nome'], 'qtd_convenios': 0}
        bancos_unicos[slug]['qtd_convenios'] += 1

out = {
    'meta': {
        'total_convenios': len(convenios),
        'total_problemas': len(problemas),
        'total_bancos_unicos': len(bancos_unicos),
        'fonte': str(XLSX.name),
    },
    'bancos_unicos': sorted(bancos_unicos.values(), key=lambda x: -x['qtd_convenios']),
    'convenios': convenios,
    'problemas': problemas,
}

OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"OK -> {OUT}")
print(f"  Convenios validos: {len(convenios)}")
print(f"  Bancos unicos:     {len(bancos_unicos)}")
print(f"  Abas problemas:    {len(problemas)}")
print(f"\nConvenios encontrados:")
for c in convenios:
    print(f"  [{c['categoria']:7s}/{c['orgao']:13s}/{c['operacao_tipo']:18s}] {c['nome']:50s} {c['qtd_bancos']} bancos")
print(f"\nTop 15 bancos mais frequentes:")
for b in sorted(bancos_unicos.values(), key=lambda x: -x['qtd_convenios'])[:15]:
    print(f"  {b['nome']:40s} {b['qtd_convenios']} convenios")
print(f"\nProblemas:")
for p in problemas:
    print(f"  - {p['sheet']}: {p['motivo']}")
