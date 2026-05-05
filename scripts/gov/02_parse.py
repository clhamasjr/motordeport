"""
Parser principal: le a planilha de governos e gera convenios.json estruturado.

Estrutura cada aba como:
  - linha 1: nome do convenio (col 2+)
  - linha 2: bancos (col 2+) - pode ter "**SUSPENSO**" no nome
  - linhas 3+: cada linha = 1 atributo (col 1 = nome do atributo, cols 2+ = valor por banco)
  - linhas com so col 1 preenchida = secao (PORTABILIDADE, CARTAO, PUBLICO ALVO)

Saida: scripts/gov/convenios.json
"""
import json, re, sys, unicodedata
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(r"C:\Users\clham\Downloads\GOVERNOS  RESUMO OPERACIONAL CONVENIOS ATUALIZADO_0_1777552913261.xlsx")
OUT = Path(__file__).parent / 'convenios.json'

# Abas indices/grupos - nao sao convenios individuais
SKIP_SHEETS = {
    'BASE','GOVERNOS','GOV PA','AM_CONVÊNIOS','TO - CONVÊNIOS','BAHIA_GOVERNOS',
    'PB_GOVERNOS','DF_GOVERNO','ES_GOVERNOS','MG_GOVERNOS','SP_GOVERNOS',
    'ELEGIVEIS','ELEGIVEIS1','GOVMG_INFORMAÇÕES OPERACIONAIS '
}

# UF -> nome estado
UF_NAMES = {
    'AC':'Acre','AL':'Alagoas','AP':'Amapá','AM':'Amazonas','BA':'Bahia',
    'CE':'Ceará','DF':'Distrito Federal','ES':'Espírito Santo','GO':'Goiás',
    'MA':'Maranhão','MT':'Mato Grosso','MS':'Mato Grosso do Sul','MG':'Minas Gerais',
    'PA':'Pará','PB':'Paraíba','PR':'Paraná','PE':'Pernambuco','PI':'Piauí',
    'RJ':'Rio de Janeiro','RN':'Rio Grande do Norte','RS':'Rio Grande do Sul',
    'RO':'Rondônia','RR':'Roraima','SC':'Santa Catarina','SP':'São Paulo',
    'SE':'Sergipe','TO':'Tocantins',
}

def slugify(s):
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^a-zA-Z0-9]+', '-', s).strip('-').lower()
    return s

def detect_uf(sheet_name):
    """Detecta UF a partir do nome da aba."""
    s = sheet_name.upper()
    # padrao mais comum: "GOV XX" ou "TJ XX" ou "MP XX" etc
    for uf in UF_NAMES:
        if re.search(rf'\b{uf}\b', s):
            return uf
    # nomes especiais
    if 'PARANA' in s: return 'PR'
    if 'RORAIMA' in s: return 'RR'
    if 'RONDONIA' in s: return 'RO'
    if 'BAHIA' in s: return 'BA'
    if 'GOIAS' in s or 'GOIÁS' in s: return 'GO'
    if 'CEARA' in s or 'CEARÁ' in s: return 'CE'
    if 'AMAPA' in s or 'AMAPÁ' in s: return 'AP'
    if 'PIAUI' in s or 'PIAUÍ' in s: return 'PI'
    if 'MARANHAO' in s or 'MARANHÃO' in s: return 'MA'
    if 'PERNAMBUCO' in s: return 'PE'
    if 'SERGIPE' in s: return 'SE'
    if 'IGEPREV' in s or 'AMAZONPREV' in s or 'AMZONPREV' in s:
        return 'TO' if 'TO' in s else 'AM'
    if 'EMBASA' in s: return 'BA'
    if 'UNICAMP' in s or 'USP' in s or 'METRO-SP' in s: return 'SP'
    if 'PREVIBANERJ' in s: return 'RJ'
    if 'C MARA DOS DEPUTADOS' in s or 'CÂMARA' in s or 'SENADO' in s or 'SUPR' in s:
        return 'DF'
    if 'MINIST' in s and ('FEDERAL' in s or 'TRABALHO' in s or 'MILITAR' in s):
        return 'DF'
    return None

def is_suspenso(banco_name):
    if not banco_name: return False
    n = banco_name.upper()
    return ('SUSPENSO' in n or 'SUPSENSO' in n or 'BLOQUEADO' in n
            or 'INATIVO' in n or 'TEMPORARIAMENTE' in n)

def clean_banco_name(n):
    if not n: return ''
    s = str(n)
    # Pega apenas a 1a linha (varios bancos tem nome multi-linha tipo "BMG\nSUSPENSO")
    s = s.split('\n')[0].split('\r')[0].strip()
    # remove ** e marcadores de status (incluindo "TEMPORARIAMENTE SUSPENSO", "INATIVO")
    s = re.sub(r'\*+[^*]*\*+', '', s)  # tudo entre ** **
    s = re.sub(r'\bTEMPORARIAMENTE\s+SUSPENS\w*\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUSPENSO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUPSENSO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bBLOQUEADO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bINATIVO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bTEMPORARIAMENTE\b', '', s, flags=re.IGNORECASE)
    s = s.replace('**','').strip(' -:')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def is_valid_banco_name(n):
    """Filtra pseudo-bancos: textos longos/explicativos que vazaram pra linha de banco."""
    if not n: return False
    s = str(n).strip()
    primeira_linha = s.split('\n')[0].split('\r')[0].strip()
    if len(primeira_linha) > 50: return False
    if len(primeira_linha) < 2: return False
    # tem ":" no meio = quase sempre instrucao explicativa, nao banco
    if ':' in primeira_linha: return False
    bad_prefixes = ('apenas','havendo','somente','obs','observacao','para ','quando ','os mesmos','dever','a partir','reconhecimento')
    if primeira_linha.lower().startswith(bad_prefixes): return False
    if not re.search(r'[a-zA-Z]', primeira_linha): return False
    # rejeita se tiver ponto final no meio (frase)
    if re.search(r'\.\s+[A-Z]', primeira_linha): return False
    return True

# ── Mapeamento de atributos pro nosso schema canonico ──────────
# Chave = slug canonico, valor = lista de regex pra match (case-insensitive, sem acentos)
ATTR_MAP = [
    ('operacoes',                ['quais operacoes', 'operacoes realizadas', 'operacoes que realiza']),
    ('valor_minimo',             ['minimo liberado novo', 'valor minimo de operacao', 'valor minimo']),
    ('margem_utilizavel',        ['margem utilizavel']),
    ('data_corte',               ['data de corte']),
    ('qtd_contratos',            ['qtde.* de contrato', 'quantidade.*contrato']),
    ('margem_seguranca',         ['^margem de seguranca$']),
    ('margem_seguranca_refin',   ['margem de seguranca p/ refin','margem de seguranca refin']),
    ('min_parcelas_refin',       ['minimo de parcelas pagas para refinanciar']),
    ('pode_agregar_margem',      ['pode agregar margem no refin', 'pode agregar margem']),
    ('reserva_margem',           ['^reserva de margem']),
    ('pode_abater_negativo',     ['pode abater o negativo']),
    ('pode_unificar_parcelas',   ['pode unificar as parcelas']),
    ('formalizacao_digital',     ['faz formalizacao digital']),
    ('limite_idade',             ['limite de idade', 'idade x valor']),
    ('limite_credito',           ['limite de credito']),
    ('liberacao_credito',        ['liberacao do credito']),
    # PORTABILIDADE
    ('port_saldo_minimo',        ['saldo minimo']),
    ('port_min_parcelas_pagas',  ['minimo de parcelas pagas']),
    ('port_taxa_minima',         ['taxa minima']),
    ('port_autorizacao_especial',['precisa de autorizacao especial']),
    ('port_pagamento_saldo',     ['pagamento de saldo']),
    ('port_doc_pagamento_saldo', ['documentacao obrigatoria para pagamento']),
    ('port_analise_quando',      ['analise antes ou depois do saldo']),
    ('port_reduz_parcela',       ['faz reducao de parcela']),
    ('port_agrega_margem',       ['^agrega margem']),
    ('port_unifica_parcelas',    ['^unifica parcelas']),
    ('port_risco',               ['risco banco ou correspondente']),
    ('port_compra_divida',       ['compra de divida']),
    ('port_o_que_precisa_compra',['o que precisa para realizar a compra']),
    # CARTAO BENEFICIO
    ('cb_limite_minimo',         ['limite minimo de credito']),
    ('cb_valor_minimo_saque',    ['valor minimo de saque']),
    ('cb_valor_minimo_margem',   ['valor minimo de margem']),
    ('cb_limite_maximo_saque',   ['limite maximo de saque']),
    ('cb_taxa_juros',            ['^taxa de juros$']),
    ('cb_seguro',                ['valor seguro prestamista']),
    ('cb_desconto_seguro',       ['desconto do seguro']),
    ('cb_analfabeto',            ['analfabeto']),
    ('cb_tarifa_emissao',        ['tarifa de emissao do cartao']),
    # PUBLICO ALVO
    ('publico_ativo',            ['ativo']),
    ('publico_aposentado',       ['aposentado']),
    ('publico_pensionista',      ['pensionista']),
]

# Secoes (linhas de cabecalho que separam blocos)
SECTION_HEADERS = ['portabilidade','cartao','cartão','publico alvo','público alvo']

def normalize_attr_name(s):
    if not s: return ''
    n = unicodedata.normalize('NFKD', str(s))
    n = ''.join(c for c in n if not unicodedata.combining(c))
    n = re.sub(r'\s+', ' ', n).strip().lower()
    return n

def match_attr_slug(attr_name):
    n = normalize_attr_name(attr_name)
    for slug, patterns in ATTR_MAP:
        for p in patterns:
            if re.search(p, n):
                return slug
    return None

def is_section_header(attr_name):
    n = normalize_attr_name(attr_name)
    return n in SECTION_HEADERS or any(n.startswith(s) for s in SECTION_HEADERS)

def parse_idade_range(text):
    """Extrai idade_min, idade_max de textos como '21 a 75 anos' ou 'De 21 anos completos 75 anos'."""
    if not text: return None, None
    s = str(text)
    # padroes comuns
    m = re.search(r'(\d{1,2})\s*(?:a|à|até|completos)?\s*(\d{2})\s*anos', s, re.IGNORECASE)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 16 <= a <= 99 and 16 <= b <= 99 and a < b:
            return a, b
    return None, None

def parse_margem(text):
    """0.35 ou 35% ou 'margem de 35%'."""
    if text is None: return None
    s = str(text).strip()
    # se for numero direto entre 0 e 1
    try:
        f = float(s.replace(',','.'))
        if 0 < f <= 1: return round(f,4)
        if 1 < f <= 100: return round(f/100,4)
    except: pass
    m = re.search(r'(\d+[\.,]?\d*)\s*%', s)
    if m:
        try: return round(float(m.group(1).replace(',','.'))/100, 4)
        except: pass
    return None

def parse_taxa(text):
    """Taxa mensal: 0.0185 ou 1.85% ou '1,85'."""
    if text is None: return None
    s = str(text).strip()
    try:
        f = float(s.replace(',','.'))
        if 0 < f < 0.1: return round(f,4)  # ja em decimal
        if 0.5 < f < 5: return round(f/100,4)  # em percentual
    except: pass
    m = re.search(r'(\d+[\.,]\d+)\s*%?', s)
    if m:
        try:
            f = float(m.group(1).replace(',','.'))
            if 0 < f < 0.1: return round(f,4)
            if 0.5 < f < 5: return round(f/100,4)
        except: pass
    return None

def parse_operacoes(text):
    """Extrai bools de Novo/Refin/Port/Cartao."""
    if not text: return {'novo':False,'refin':False,'port':False,'cartao':False}
    s = str(text).upper()
    return {
        'novo':   bool(re.search(r'\bNOVO\b', s)),
        'refin':  bool(re.search(r'\bREFIN', s)),
        'port':   bool(re.search(r'\bPORT', s)),
        'cartao': bool(re.search(r'CART[ÃA]O|\bCB\b', s)),
    }

# ─── PARSE ────────────────────────────────────────────────────
wb = openpyxl.load_workbook(XLSX, data_only=True)
convenios = []
problemas = []

for sn in wb.sheetnames:
    if sn in SKIP_SHEETS: continue
    ws = wb[sn]
    rows = []
    # le ate 80 linhas, capturando colunas A-AZ no max
    max_cols = min(ws.max_column or 1, 30)
    if max_cols < 2:
        problemas.append({'sheet': sn, 'motivo': 'sem colunas'})
        continue
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), max_col=max_cols, values_only=True):
        rows.append([v for v in row])
    # remove linhas totalmente vazias
    rows = [r for r in rows if any(v not in (None, '') and str(v).strip() for v in r)]
    if len(rows) < 3:
        problemas.append({'sheet': sn, 'motivo': f'so {len(rows)} linhas com dados'})
        continue

    # Linha 1: nome do convenio (procura na col 2+ a primeira celula nao vazia)
    nome_conv = None
    for v in rows[0][1:]:
        if v and str(v).strip():
            nome_conv = str(v).strip()
            break
    if not nome_conv: nome_conv = sn.strip()

    # Linha 2: bancos
    bancos_header = []
    for i, v in enumerate(rows[1][1:], start=1):
        if v and str(v).strip() and is_valid_banco_name(v):
            bancos_header.append({'col': i, 'nome_raw': str(v).strip()})
    if not bancos_header:
        problemas.append({'sheet': sn, 'motivo': 'sem bancos identificados'})
        continue

    # Linhas 3+: atributos
    atributos = []  # cada item = {slug, label_raw, valores: {col: text}}
    secao_atual = 'principal'
    for r in rows[2:]:
        attr_label = r[0]
        if not attr_label or not str(attr_label).strip(): continue
        attr_label = str(attr_label).strip()
        valores_cols = {i: ('' if v is None else str(v).strip()) for i, v in enumerate(r[1:], start=1)}
        # detecta secao (linha onde so col A tem texto)
        if is_section_header(attr_label) and not any(valores_cols.values()):
            n = normalize_attr_name(attr_label)
            if 'portabilidade' in n: secao_atual = 'portabilidade'
            elif 'cartao' in n or 'cartão' in n: secao_atual = 'cartao'
            elif 'publico' in n or 'público' in n: secao_atual = 'publico_alvo'
            continue
        slug = match_attr_slug(attr_label)
        atributos.append({
            'slug': slug, 'label_raw': attr_label,
            'secao': secao_atual, 'valores': valores_cols
        })

    # Monta os bancos com seus atributos
    bancos_out = []
    for b in bancos_header:
        col = b['col']
        nome_raw = b['nome_raw']
        suspenso = is_suspenso(nome_raw)
        nome_clean = clean_banco_name(nome_raw)
        # Skip "headers que nao sao bancos" (ex: 'PORTABILIDADE' acidentalmente na linha 2)
        if not nome_clean or len(nome_clean) < 2: continue

        atrib_canonicos = {}   # slug -> texto raw
        atrib_brutos = []      # [{label, valor, secao}]
        for a in atributos:
            valor = a['valores'].get(col, '').strip()
            if not valor: continue
            atrib_brutos.append({
                'label': a['label_raw'], 'valor': valor, 'secao': a['secao']
            })
            if a['slug']:
                atrib_canonicos[a['slug']] = valor

        # parse campos numericos
        operacoes = parse_operacoes(atrib_canonicos.get('operacoes',''))
        idade_min, idade_max = parse_idade_range(
            atrib_canonicos.get('limite_idade','') + ' ' + atrib_canonicos.get('limite_credito','')
        )
        margem = parse_margem(atrib_canonicos.get('margem_utilizavel'))
        taxa = parse_taxa(atrib_canonicos.get('port_taxa_minima'))

        bancos_out.append({
            'nome': nome_clean,
            'nome_raw': nome_raw,
            'slug': slugify(nome_clean),
            'suspenso': suspenso,
            'operacoes': operacoes,
            'idade_min': idade_min,
            'idade_max': idade_max,
            'margem_utilizavel': margem,
            'taxa_minima_port': taxa,
            'atributos': atrib_canonicos,
            'atributos_brutos': atrib_brutos,
        })

    if not bancos_out:
        problemas.append({'sheet': sn, 'motivo': 'sem bancos validos apos limpeza'})
        continue

    uf = detect_uf(sn)
    convenios.append({
        'sheet': sn,
        'slug': slugify(sn),
        'nome': nome_conv,
        'uf': uf,
        'estado_nome': UF_NAMES.get(uf) if uf else None,
        'qtd_bancos': len(bancos_out),
        'qtd_atributos': len(atributos),
        'bancos': bancos_out,
    })

# ─── Agrega bancos unicos pra criar tabela mestre ──────────────
bancos_unicos = {}
for c in convenios:
    for b in c['bancos']:
        slug = b['slug']
        if slug not in bancos_unicos:
            bancos_unicos[slug] = {'slug': slug, 'nome': b['nome'], 'qtd_convenios': 0}
        bancos_unicos[slug]['qtd_convenios'] += 1

out = {
    'meta': {
        'total_convenios': len(convenios),
        'total_problemas': len(problemas),
        'total_bancos_unicos': len(bancos_unicos),
        'fonte': str(XLSX.name),
    },
    'bancos_unicos': sorted(bancos_unicos.values(), key=lambda x: -x['qtd_convenios']),
    'convenios': convenios,
    'problemas': problemas,
}

OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"OK -> {OUT}")
print(f"  Convenios validos: {len(convenios)}")
print(f"  Bancos unicos:     {len(bancos_unicos)}")
print(f"  Abas problemas:    {len(problemas)}")
print(f"\nTop 10 bancos mais frequentes:")
for b in sorted(bancos_unicos.values(), key=lambda x: -x['qtd_convenios'])[:10]:
    print(f"  {b['nome']:40s} {b['qtd_convenios']} convenios")
print(f"\nProblemas:")
for p in problemas[:20]:
    print(f"  - {p['sheet']}: {p['motivo']}")
