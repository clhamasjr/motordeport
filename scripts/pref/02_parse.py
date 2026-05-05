"""
Parser principal: le a planilha de PREFEITURAS e gera convenios.json estruturado.

Estrutura igual ao gov:
  - linha 1: nome do convenio (col 2+)
  - linha 2: bancos (col 2+) - pode ter "**SUSPENSO**" no nome
  - linhas 3+: cada linha = 1 atributo (col 1 = nome do atributo, cols 2+ = valor por banco)
  - linhas com so col 1 preenchida = secao (PORTABILIDADE, CARTAO, PUBLICO ALVO)

DIFERENCAS pra GOV:
  - Planilha tem ~622 abas (vs ~120 do gov), 1 por municipio/instituto/CB
  - Detecao de UF por CONTEXTO: a planilha agrupa abas por UF — abas pequenas (1x1
    ou indices) tipo "AC", "AL", "AM", "PE" servem como SEPARADORES entre as UFs
  - Extrai campo `municipio` do nome da aba
  - Extrai campo `tipo` ('prefeitura' | 'instituto_previdencia' | 'cartao_beneficio')

Saida: scripts/pref/convenios.json
"""
import json, re, sys, unicodedata
from pathlib import Path
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

XLSX = Path(r"C:\Users\clham\Downloads\PREFEITURAS - RESUMO OPERACIONAL CONVENIOS_0_1777485648411.xlsx")
OUT = Path(__file__).parent / 'convenios.json'

# ── Abas que sao indices/grupos/separadores - nao sao convenios ──
SKIP_SHEETS = {
    'BASE','PREFEITURAS','PREFEITURAS PB','PREFEITURAS PI','PREFEITURAS ANAPOLIS',
    'PREFEITURAS MACAE','PREFEITURAS CARAGUATATUBA','PREFEITURAS FLORIANOPOLIS',
    'PREFEITURAS JARAGUA DO SUL ','PREFEITURA JOINVILLE','PREFEITURA LONDRINA',
    'PREFEITURA CURITIBA','PREFEITURA CASCAVEL','PREFEITURA DE PONTA GROSSA',
    'PREFEITURA DE PALHOÇA','PREFEITURA DE SÃO PAULO','PREFEITURA DE OSACO',
    'PREFEITURA NATAL','PREFEITURA POÁ','PREFEITURA MARILIA ','PREFEITURA BAURU',
    'PREFEITURA DE LONDRINA- CAAPSML',  # essa eh subcategoria
    'CIDADES_SP','PREF_MG','PREF_DIVINOPOLIS','PREF_UBERABA','PREF MA',
    'PREF. JUIZ','SITUAÇÃO ACEITAS E NÃO ACEITAS',
    # UFs sozinhas (separadores)
    'AC','AL','AM','AMAPÁ','PA','TO','BAHIA','CEARÁ','PE','RN','SE','GO','MT','MS','ES',
    'RJ','PR','RS','SC','RONDONIA','BOA VISTA',
}

# ── UF -> nome estado ──
UF_NAMES = {
    'AC':'Acre','AL':'Alagoas','AP':'Amapá','AM':'Amazonas','BA':'Bahia',
    'CE':'Ceará','DF':'Distrito Federal','ES':'Espírito Santo','GO':'Goiás',
    'MA':'Maranhão','MT':'Mato Grosso','MS':'Mato Grosso do Sul','MG':'Minas Gerais',
    'PA':'Pará','PB':'Paraíba','PR':'Paraná','PE':'Pernambuco','PI':'Piauí',
    'RJ':'Rio de Janeiro','RN':'Rio Grande do Norte','RS':'Rio Grande do Sul',
    'RO':'Rondônia','RR':'Roraima','SC':'Santa Catarina','SP':'São Paulo',
    'SE':'Sergipe','TO':'Tocantins',
}

# Abas-separador que MARCAM inicio de UF (a partir delas, todas as abas seguintes
# sao da UF marcada, ate encontrar outra abas-separador)
UF_SEPARATORS = {
    'AC':'AC','AL':'AL','AM':'AM','AMAPÁ':'AP','AP':'AP','PA':'PA','TO':'TO',
    'BAHIA':'BA','BA':'BA','CEARÁ':'CE','CE':'CE','PREF MA':'MA','MA':'MA',
    'PREFEITURAS PB':'PB','PB':'PB','PREFEITURAS PI':'PI','PI':'PI',
    'PE':'PE','RN':'RN','SE':'SE','GO':'GO','MT':'MT','MS':'MS','ES':'ES',
    'PREF_MG':'MG','MG':'MG','RJ':'RJ','CIDADES_SP':'SP','SP':'SP',
    'PR':'PR','RS':'RS','SC':'SC','RONDONIA':'RO','BOA VISTA':'RR','RR':'RR',
}

def slugify(s):
    s = unicodedata.normalize('NFKD', str(s))
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r'[^a-zA-Z0-9]+', '-', s).strip('-').lower()
    return s

def clean_text(s):
    """Remove acentos e normaliza."""
    n = unicodedata.normalize('NFKD', str(s))
    n = ''.join(c for c in n if not unicodedata.combining(c))
    return n.strip()

# ── Detecta tipo de aba (pref / instituto / cb) ──
def detect_tipo(sheet_name):
    s = clean_text(sheet_name).upper()
    if re.search(r'\bCB\b|CART[ÃA]O|CART\.?\s*BEN[EF]', s):
        return 'cartao_beneficio'
    # institutos de previdencia
    inst_keywords = ['PREV','IPM','IPS','IPSEM','IPMDC','IPC','IPVV','IPAMV','IPREM','IPREJUN',
                     'IPREVILLE','IPREVGUARULHOS','IPREMU','IPSERV','MAPRO','JFPREV','PROCON',
                     'TRT','MACAEPREV','TEREPREV','SEPREM','EMBUPREV','CAMPREV','FUNPREV',
                     'CARAGUAPREV','CMT','CMC','PREVIBAM','MARINGAPREV','MARINGÁPREV','FOZPREV',
                     'CAIXA','ASCMC','HSPM','AHM','OMSS','SAME','ISSEM','ISSM','ISSA','SJRPPREV',
                     'MAPRO','DEMAE','DAERP','CODAU','COMCAP','PREVMOC','DIVIPREV','PREVCEL',
                     'PREVIPALMAS','MANAUSPREV','MACAPA PREV','MACAPAPREV','GURUPI PREV','PREVIBANERJ',
                     'IGEPREV','IPASG','IPREMB','CAMPREV','ITAJAÍ','TANGARÁ DA SERRA','CACERES',
                     'CALDASPREV','PREVCAL','IPAM','CUIABAPREV','VALPARAISO',
                     'CAXIASPREV','PROCURADORIA','CEARÁ-MIRIAM','CAUCAIA',
                     'INST','PREV','JFPREV','DEMAE','DEPMSA','CODAU','OMSS']
    for k in inst_keywords:
        if k in s and 'PREF' not in s.split()[0]:
            return 'instituto_previdencia'
    if s.startswith('PREF') or 'PREFEITURA' in s:
        return 'prefeitura'
    return 'outro'

# ── Extrai municipio do nome da aba ──
def extract_municipio(sheet_name):
    """Tira prefixos tipo PREF, CB, CART.BENF, INST PREV — sobra o municipio."""
    if not sheet_name: return None
    s = sheet_name.strip()
    # remove sigla UF no fim
    s = re.sub(r'\s*[-–]?\s*\b(AC|AL|AM|AP|BA|CE|DF|ES|GO|MA|MT|MS|MG|PA|PB|PR|PE|PI|RJ|RN|RS|RO|RR|SC|SP|SE|TO)\b\s*$', '', s, flags=re.IGNORECASE)
    # remove prefixos comuns
    prefixos = [
        r'^PREFEITURA\s+(DE|DA|DO|DAS|DOS)?\s*',
        r'^PREFEITURAS\s+',
        r'^PREF\.?\s+(DE|DA|DO|DAS|DOS)?\s*',
        r'^Pref\.?\s+(de|da|do|das|dos)?\s*',
        r'^CART\.?\s*BEN[EFI]+\.?\s+(DE|DA|DO|DAS|DOS)?\s*',
        r'^CARTÃO\s+BEN(E|EFI)+CIO\s+(DE|DA|DO|DAS|DOS)?\s*',
        r'^CARTAO\s+BEN[EFI]+CIO\s+(DE|DA|DO|DAS|DOS)?\s*',
        r'^CB\s+(DE|DA|DO|DAS|DOS|PREF\.?)?\s*',
        r'^INST\.?\s*PREV\.?\s+(DE|DA|DO|DAS|DOS|S\.J\.)?\s*',
        r'^INST\.?\s*SOCI\.?\s*SERV\s+',
        r'^INSTITUTO\s+(IPSEM|IPMDC|PREV)?\s*',
    ]
    for p in prefixos:
        s2 = re.sub(p, '', s, flags=re.IGNORECASE).strip()
        if s2 and s2 != s:
            s = s2
            break
    # sufixos comuns
    sufixos = [r'\s+(CB|PREV|CARTÃO\s+CONSIGNA)$', r'\s+CARTÃO\s+BENEF\.?$', r'\s+CART\..*$']
    for sf in sufixos:
        s = re.sub(sf, '', s, flags=re.IGNORECASE).strip()
    # remove pontuacao final
    s = s.strip(' .,-:;')
    return s or sheet_name.strip()

def is_suspenso(banco_name):
    if not banco_name: return False
    n = clean_text(banco_name).upper()
    return ('SUSPENSO' in n or 'SUPSENSO' in n or 'BLOQUEADO' in n
            or 'INATIVO' in n or 'TEMPORARIAMENTE' in n)

def clean_banco_name(n):
    if not n: return ''
    s = str(n)
    s = s.split('\n')[0].split('\r')[0].strip()
    s = re.sub(r'\*+[^*]*\*+', '', s)
    s = re.sub(r'\bTEMPORARIAMENTE\s+SUSPENS\w*\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUSPENSO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bSUPSENSO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bBLOQUEADO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bINATIVO\b', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\bTEMPORARIAMENTE\b', '', s, flags=re.IGNORECASE)
    s = s.replace('**','').strip(' -:')
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def is_valid_banco_name(n):
    if not n: return False
    s = str(n).strip()
    primeira_linha = s.split('\n')[0].split('\r')[0].strip()
    if len(primeira_linha) > 50: return False
    if len(primeira_linha) < 2: return False
    if ':' in primeira_linha: return False
    bad_prefixes = ('apenas','havendo','somente','obs','observacao','para ','quando ','os mesmos','dever','a partir','reconhecimento')
    if primeira_linha.lower().startswith(bad_prefixes): return False
    if not re.search(r'[a-zA-Z]', primeira_linha): return False
    if re.search(r'\.\s+[A-Z]', primeira_linha): return False
    return True

# ── Mapeamento de atributos canonicos (mesmo do gov) ──────────
ATTR_MAP = [
    ('operacoes',                ['quais operacoes', 'operacoes realizadas', 'operacoes que realiza']),
    ('valor_minimo',             ['minimo liberado novo', 'valor minimo de operacao', 'valor minimo']),
    ('margem_utilizavel',        ['margem utilizavel']),
    ('data_corte',               ['data de corte']),
    ('qtd_contratos',            ['qtde.* de contrato', 'quantidade.*contrato']),
    ('margem_seguranca',         ['^margem de seguranca$']),
    ('margem_seguranca_refin',   ['margem de seguranca p/ refin','margem de seguranca refin']),
    ('min_parcelas_refin',       ['minimo de parcelas pagas para refinanciar']),
    ('pode_agregar_margem',      ['pode agregar margem no refin', 'pode agregar margem']),
    ('reserva_margem',           ['^reserva de margem']),
    ('pode_abater_negativo',     ['pode abater o negativo']),
    ('pode_unificar_parcelas',   ['pode unificar as parcelas']),
    ('formalizacao_digital',     ['faz formalizacao digital']),
    ('limite_idade',             ['limite de idade', 'idade x valor']),
    ('limite_credito',           ['limite de credito']),
    ('liberacao_credito',        ['liberacao do credito']),
    ('port_saldo_minimo',        ['saldo minimo']),
    ('port_min_parcelas_pagas',  ['minimo de parcelas pagas']),
    ('port_taxa_minima',         ['taxa minima']),
    ('port_autorizacao_especial',['precisa de autorizacao especial']),
    ('port_pagamento_saldo',     ['pagamento de saldo']),
    ('port_doc_pagamento_saldo', ['documentacao obrigatoria para pagamento']),
    ('port_analise_quando',      ['analise antes ou depois do saldo']),
    ('port_reduz_parcela',       ['faz reducao de parcela']),
    ('port_agrega_margem',       ['^agrega margem']),
    ('port_unifica_parcelas',    ['^unifica parcelas']),
    ('port_risco',               ['risco banco ou correspondente']),
    ('port_compra_divida',       ['compra de divida']),
    ('port_o_que_precisa_compra',['o que precisa para realizar a compra']),
    ('cb_limite_minimo',         ['limite minimo de credito']),
    ('cb_valor_minimo_saque',    ['valor minimo de saque']),
    ('cb_valor_minimo_margem',   ['valor minimo de margem']),
    ('cb_limite_maximo_saque',   ['limite maximo de saque']),
    ('cb_taxa_juros',            ['^taxa de juros$']),
    ('cb_seguro',                ['valor seguro prestamista']),
    ('cb_desconto_seguro',       ['desconto do seguro']),
    ('cb_analfabeto',            ['analfabeto']),
    ('cb_tarifa_emissao',        ['tarifa de emissao do cartao']),
    ('publico_ativo',            ['ativo']),
    ('publico_aposentado',       ['aposentado']),
    ('publico_pensionista',      ['pensionista']),
]

SECTION_HEADERS = ['portabilidade','cartao','cartão','publico alvo','público alvo']

def normalize_attr_name(s):
    if not s: return ''
    n = unicodedata.normalize('NFKD', str(s))
    n = ''.join(c for c in n if not unicodedata.combining(c))
    n = re.sub(r'\s+', ' ', n).strip().lower()
    return n

def match_attr_slug(attr_name):
    n = normalize_attr_name(attr_name)
    for slug, patterns in ATTR_MAP:
        for p in patterns:
            if re.search(p, n):
                return slug
    return None

def is_section_header(attr_name):
    n = normalize_attr_name(attr_name)
    return n in SECTION_HEADERS or any(n.startswith(s) for s in SECTION_HEADERS)

def parse_idade_range(text):
    if not text: return None, None
    s = str(text)
    m = re.search(r'(\d{1,2})\s*(?:a|à|até|completos)?\s*(\d{2})\s*anos', s, re.IGNORECASE)
    if m:
        a, b = int(m.group(1)), int(m.group(2))
        if 16 <= a <= 99 and 16 <= b <= 99 and a < b:
            return a, b
    return None, None

def parse_margem(text):
    if text is None: return None
    s = str(text).strip()
    try:
        f = float(s.replace(',','.'))
        if 0 < f <= 1: return round(f,4)
        if 1 < f <= 100: return round(f/100,4)
    except: pass
    m = re.search(r'(\d+[\.,]?\d*)\s*%', s)
    if m:
        try: return round(float(m.group(1).replace(',','.'))/100, 4)
        except: pass
    return None

def parse_taxa(text):
    if text is None: return None
    s = str(text).strip()
    try:
        f = float(s.replace(',','.'))
        if 0 < f < 0.1: return round(f,4)
        if 0.5 < f < 5: return round(f/100,4)
    except: pass
    m = re.search(r'(\d+[\.,]\d+)\s*%?', s)
    if m:
        try:
            f = float(m.group(1).replace(',','.'))
            if 0 < f < 0.1: return round(f,4)
            if 0.5 < f < 5: return round(f/100,4)
        except: pass
    return None

def parse_operacoes(text):
    if not text: return {'novo':False,'refin':False,'port':False,'cartao':False}
    s = str(text).upper()
    return {
        'novo':   bool(re.search(r'\bNOVO\b', s)),
        'refin':  bool(re.search(r'\bREFIN', s)),
        'port':   bool(re.search(r'\bPORT', s)),
        'cartao': bool(re.search(r'CART[ÃA]O|\bCB\b', s)),
    }

# ─── PARSE ────────────────────────────────────────────────────
print(f'Lendo {XLSX.name}...')
wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
convenios = []
problemas = []

# UF "atual" — vai sendo atualizada conforme percorre as abas em ORDEM
current_uf = None

for sn in wb.sheetnames:
    sn_clean = sn.strip()
    sn_norm = sn_clean.upper().strip()

    # ── Eh separador de UF? ──
    if sn_norm in UF_SEPARATORS:
        current_uf = UF_SEPARATORS[sn_norm]
        # tambem skipar (eh so cabecalho)
        continue
    if sn_clean in SKIP_SHEETS:
        continue

    ws = wb[sn]
    # evita abas vazias (1x1)
    if (ws.max_row or 0) <= 2 and (ws.max_column or 0) <= 2:
        continue

    rows = []
    max_cols = min(ws.max_column or 1, 30)
    if max_cols < 2:
        problemas.append({'sheet': sn, 'motivo': 'sem colunas'})
        continue
    max_rows = min(ws.max_row or 0, 80)
    for row in ws.iter_rows(min_row=1, max_row=max_rows, max_col=max_cols, values_only=True):
        rows.append([v for v in row])
    rows = [r for r in rows if any(v not in (None, '') and str(v).strip() for v in r)]
    if len(rows) < 3:
        problemas.append({'sheet': sn, 'motivo': f'so {len(rows)} linhas com dados'})
        continue

    # Linha 1: nome do convenio
    nome_conv = None
    for v in rows[0][1:]:
        if v and str(v).strip():
            nome_conv = str(v).strip()
            break
    if not nome_conv: nome_conv = sn.strip()

    # Linha 2: bancos
    bancos_header = []
    for i, v in enumerate(rows[1][1:], start=1):
        if v and str(v).strip() and is_valid_banco_name(v):
            bancos_header.append({'col': i, 'nome_raw': str(v).strip()})
    if not bancos_header:
        problemas.append({'sheet': sn, 'motivo': 'sem bancos identificados'})
        continue

    # Linhas 3+: atributos
    atributos = []
    secao_atual = 'principal'
    for r in rows[2:]:
        attr_label = r[0]
        if not attr_label or not str(attr_label).strip(): continue
        attr_label = str(attr_label).strip()
        valores_cols = {i: ('' if v is None else str(v).strip()) for i, v in enumerate(r[1:], start=1)}
        if is_section_header(attr_label) and not any(valores_cols.values()):
            n = normalize_attr_name(attr_label)
            if 'portabilidade' in n: secao_atual = 'portabilidade'
            elif 'cartao' in n or 'cartão' in n: secao_atual = 'cartao'
            elif 'publico' in n or 'público' in n: secao_atual = 'publico_alvo'
            continue
        slug = match_attr_slug(attr_label)
        atributos.append({
            'slug': slug, 'label_raw': attr_label,
            'secao': secao_atual, 'valores': valores_cols
        })

    bancos_out = []
    for b in bancos_header:
        col = b['col']
        nome_raw = b['nome_raw']
        suspenso = is_suspenso(nome_raw)
        nome_clean = clean_banco_name(nome_raw)
        if not nome_clean or len(nome_clean) < 2: continue

        atrib_canonicos = {}
        atrib_brutos = []
        for a in atributos:
            valor = a['valores'].get(col, '').strip()
            if not valor: continue
            atrib_brutos.append({
                'label': a['label_raw'], 'valor': valor, 'secao': a['secao']
            })
            if a['slug']:
                atrib_canonicos[a['slug']] = valor

        operacoes = parse_operacoes(atrib_canonicos.get('operacoes',''))
        idade_min, idade_max = parse_idade_range(
            atrib_canonicos.get('limite_idade','') + ' ' + atrib_canonicos.get('limite_credito','')
        )
        margem = parse_margem(atrib_canonicos.get('margem_utilizavel'))
        taxa = parse_taxa(atrib_canonicos.get('port_taxa_minima'))

        bancos_out.append({
            'nome': nome_clean,
            'nome_raw': nome_raw,
            'slug': slugify(nome_clean),
            'suspenso': suspenso,
            'operacoes': operacoes,
            'idade_min': idade_min,
            'idade_max': idade_max,
            'margem_utilizavel': margem,
            'taxa_minima_port': taxa,
            'atributos': atrib_canonicos,
            'atributos_brutos': atrib_brutos,
        })

    if not bancos_out:
        problemas.append({'sheet': sn, 'motivo': 'sem bancos validos apos limpeza'})
        continue

    municipio = extract_municipio(sn)
    tipo = detect_tipo(sn)
    convenios.append({
        'sheet': sn,
        'slug': slugify(sn),
        'nome': nome_conv,
        'uf': current_uf,
        'estado_nome': UF_NAMES.get(current_uf) if current_uf else None,
        'municipio': municipio,
        'tipo': tipo,
        'qtd_bancos': len(bancos_out),
        'qtd_atributos': len(atributos),
        'bancos': bancos_out,
    })

# ─── Agrega bancos unicos ─────────────────────────────────────
bancos_unicos = {}
for c in convenios:
    for b in c['bancos']:
        slug = b['slug']
        if slug not in bancos_unicos:
            bancos_unicos[slug] = {'slug': slug, 'nome': b['nome'], 'qtd_convenios': 0}
        bancos_unicos[slug]['qtd_convenios'] += 1

out = {
    'meta': {
        'total_convenios': len(convenios),
        'total_problemas': len(problemas),
        'total_bancos_unicos': len(bancos_unicos),
        'fonte': str(XLSX.name),
    },
    'bancos_unicos': sorted(bancos_unicos.values(), key=lambda x: -x['qtd_convenios']),
    'convenios': convenios,
    'problemas': problemas,
}

OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
print(f"OK -> {OUT}")
print(f"  Convenios validos: {len(convenios)}")
print(f"  Bancos unicos:     {len(bancos_unicos)}")
print(f"  Abas problemas:    {len(problemas)}")
print(f"\nDistribuicao por UF:")
por_uf = {}
for c in convenios:
    k = c['uf'] or 'SEM_UF'
    por_uf[k] = por_uf.get(k, 0) + 1
for uf in sorted(por_uf, key=lambda k: -por_uf[k]):
    print(f"  {uf}: {por_uf[uf]}")
print(f"\nDistribuicao por tipo:")
por_tipo = {}
for c in convenios:
    por_tipo[c['tipo']] = por_tipo.get(c['tipo'], 0) + 1
for t, n in sorted(por_tipo.items(), key=lambda x: -x[1]):
    print(f"  {t}: {n}")
print(f"\nTop 15 bancos mais frequentes:")
for b in sorted(bancos_unicos.values(), key=lambda x: -x['qtd_convenios'])[:15]:
    print(f"  {b['nome']:40s} {b['qtd_convenios']} convenios")
print(f"\nProblemas (primeiros 20):")
for p in problemas[:20]:
    print(f"  - {p['sheet']}: {p['motivo']}")
